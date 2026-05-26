"""
Wooffy Breed Article Generator
Usage:
  python generate.py golden-retriever           # generate HTML only -> output/
  python generate.py --all                      # generate all breed-data/*.json
  python generate.py golden-retriever --publish # generate + publish to Shopify
  python generate.py golden-retriever --update  # update existing article (by handle)
  python generate.py --get-blog-id              # list all blogs to find SHOPIFY_BLOG_ID
  python generate.py --update-articles-md       # regenerate ARTICLES.md from breed-data/
"""

import json
import os
import re
import sys
import glob
import urllib.request
import urllib.error
from datetime import datetime

ROOT_DIR       = os.path.join(os.path.dirname(__file__), "..")
BREED_DATA_DIR = os.path.join(ROOT_DIR, "breed-data")
OUTPUT_DIR     = os.path.join(ROOT_DIR, "output")
LOG_PATH       = os.path.join(ROOT_DIR, "published_log.json")
XLSX_PATH      = os.path.join(ROOT_DIR, "ARTICLES.xlsx")
ENV_PATH       = os.path.join(ROOT_DIR, ".env")

BLOG_BASE    = "/blogs/dog-breeds"
ARTICLES_PATH = os.path.join(ROOT_DIR, "ARTICLES.md")


# ── Load .env ─────────────────────────────────────────────────────────────────
def load_env():
    env = {}
    if os.path.exists(ENV_PATH):
        with open(ENV_PATH) as f:
            for line in f:
                line = line.strip()
                if line and "=" in line and not line.startswith("#"):
                    k, v = line.split("=", 1)
                    env[k.strip()] = v.strip()
    return env

_env = load_env()
SHOPIFY_STORE   = _env.get("SHOPIFY_STORE",   "your-store.myshopify.com")
SHOPIFY_TOKEN   = _env.get("SHOPIFY_TOKEN",   "")
SHOPIFY_BLOG_ID = int(_env.get("SHOPIFY_BLOG_ID", "0") or "0")
SHOPIFY_DOG_NUTRITION_BLOG_ID = int(
    _env.get("SHOPIFY_DOG_NUTRITION_BLOG_ID", "0") or "0"
)

# blog_handle -> blog id. Extend as more blogs come under autofix.
# Env var names follow the pattern SHOPIFY_<HANDLE>_BLOG_ID (SHOPIFY_BLOG_ID
# is kept as the legacy alias for the dog-breeds blog).
BLOG_IDS = {
    "dog-breeds":    SHOPIFY_BLOG_ID,
    "dog-nutrition": SHOPIFY_DOG_NUTRITION_BLOG_ID,
}


def _resolve_blog(breed):
    """Return (blog_handle, blog_id) for this article. Default = dog-breeds
    so every existing breed-data file routes correctly with no migration."""
    handle = (breed.get("meta") or {}).get("blog_handle") or "dog-breeds"
    blog_id = BLOG_IDS.get(handle, 0)
    if not blog_id:
        env_var = "SHOPIFY_" + handle.upper().replace("-", "_") + "_BLOG_ID"
        raise RuntimeError(
            f"No blog id configured for blog_handle={handle!r}. "
            f"Set {env_var} in .env."
        )
    return handle, blog_id
# ─────────────────────────────────────────────────────────────────────────────


def add_link_targets(html):
    """Add target="_blank" rel="noopener noreferrer" to all links except # anchors."""
    def replacer(m):
        tag = m.group(0)
        href_match = re.search(r'href="([^"]*)"', tag)
        if href_match and href_match.group(1).startswith('#'):
            return tag
        if 'target=' in tag:
            return tag
        return tag.replace('<a ', '<a target="_blank" rel="noopener noreferrer" ', 1)
    return re.sub(r'<a [^>]+>', replacer, html)


def load_breed(slug):
    path = os.path.join(BREED_DATA_DIR, f"{slug}.json")
    if not os.path.exists(path):
        raise FileNotFoundError(f"No breed data found: {path}")
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def build_stats_grid(stats):
    cards = ""
    for key, s in stats.items():
        cards += f"""
    <div style="background:#ffffff;border-radius:8px;padding:14px 10px;text-align:center;border:1px solid rgba(0,0,0,0.1);">
      <div style="font-size:1.2em;margin-bottom:6px;">{s['emoji']}</div>
      <div style="font-family:'figmaMono','SF Mono',monospace;font-size:0.58em;font-weight:400;letter-spacing:0.54px;text-transform:uppercase;color:rgba(0,0,0,0.5);margin-bottom:4px;">{s['label']}</div>
      <div style="font-size:0.85em;font-weight:700;color:#000000;letter-spacing:-0.14px;">{s['value']}</div>
    </div>"""
    return f'<div style="display:grid;grid-template-columns:repeat(4,1fr);gap:8px;">{cards}\n  </div>'


def build_toc(sections, breed):
    # Use the article's actual section insertion order (Python 3.7+ preserves it),
    # plus always append faq at the end if the article has FAQ entries.
    # This handles main breed articles, roundups, and comparison articles uniformly.
    has_faq = bool(breed.get("faq"))
    ordered = list(sections.keys())
    if has_faq and "faq" not in ordered:
        ordered.append("faq")
    fallbacks = {
        "right_for_you": sections.get("right_for_you", {}).get("heading", "Is This Breed Right for You?"),
        "faq":           "FAQ",
    }
    items = ""
    for key in ordered:
        if key == "faq":
            if not has_faq:
                continue
            sec = {}
        else:
            sec = sections.get(key, {})
            if not isinstance(sec, dict):
                continue
            # Skip empty html sections (right_for_you may have good_fit instead of html)
            if key != "right_for_you" and not sec.get("html"):
                continue
            if key == "right_for_you" and not sec.get("good_fit") and not sec.get("html"):
                continue
        anchor = key.replace("_", "-")
        heading = sec.get("heading", "") if isinstance(sec, dict) else ""
        label = heading or fallbacks.get(key, key)
        label = label.replace("<br>", " ").replace("<br/>", " ")
        items += f'    <li><a href="#{anchor}">{label}</a></li>\n'
    return f"""<aside class="wooffy-toc">
  <p style="font-family:'figmaMono','SF Mono',monospace;font-size:0.7em;font-weight:400;letter-spacing:0.54px;text-transform:uppercase;color:rgba(0,0,0,0.45);margin:0 0 12px 0;padding-left:12px;">In This Guide</p>
  <ol>
{items}  </ol>
</aside>"""


def build_section(key, data, breed_name):
    anchor = key.replace("_", "-")
    label  = data.get("label", "")
    heading = data.get("heading", "")
    html   = data.get("html", "")

    label_html    = f'<p style="font-family:\'figmaMono\',\'SF Mono\',monospace;font-size:0.7em;font-weight:400;letter-spacing:0.54px;text-transform:uppercase;color:rgba(0,0,0,0.45);margin:0 0 8px 0;">{label}</p>' if label else ""
    heading_html  = f'<h2 id="{anchor}" style="font-size:1.5em;font-weight:700;color:#1a1a1a;margin:0 0 16px 0;line-height:1.2;">{heading}</h2>' if heading else ""

    return f"""<div style="margin-bottom:48px;padding-top:32px;border-top:1px solid #e8e8e8;">
  {label_html}
  {heading_html}
  {html}
</div>"""


def build_right_for_you(data):
    good_rows = "".join(
        f'<tr style="border-bottom:1px solid #e8e8e8;"><td style="padding:10px 14px;color:#6b7177;background:#f8f8f8;">{item}</td>'
        f'<td style="padding:10px 14px;color:#6b7177;background:#f8f8f8;">{data["not_fit"][i] if i < len(data["not_fit"]) else ""}</td></tr>'
        for i, item in enumerate(data["good_fit"])
    )
    return f"""<div style="margin-bottom:48px;padding-top:32px;border-top:1px solid #e8e8e8;">
  <p style="font-family:'figmaMono','SF Mono',monospace;font-size:0.7em;font-weight:400;letter-spacing:0.54px;text-transform:uppercase;color:rgba(0,0,0,0.45);margin:0 0 8px 0;">{data.get('label','Fit Assessment')}</p>
  <h2 id="right-for-you" style="font-size:1.5em;font-weight:700;color:#1a1a1a;margin:0 0 20px 0;line-height:1.2;">{data.get('heading','')}</h2>
  <table style="width:100%;font-size:0.88em;">
    <thead>
      <tr>
        <th style="padding:12px 14px;text-align:left;font-weight:700;color:#006b3c;background:#f0faf4;border-bottom:2px solid #a8d5b5;width:50%;">Great fit if you...</th>
        <th style="padding:12px 14px;text-align:left;font-weight:700;color:#c62828;background:#fff5f5;border-bottom:2px solid #ffcdd2;width:50%;">Not the best fit if you...</th>
      </tr>
    </thead>
    <tbody>
      {good_rows}
    </tbody>
  </table>
</div>"""


def build_fun_facts(facts):
    cards = ""
    for f in facts:
        cards += f"""
    <div style="background:#ffffff;border:1.5px solid #c8c8c8;border-radius:20px;padding:28px 24px;">
      <div style="font-size:2em;margin-bottom:14px;">{f['emoji']}</div>
      <div style="font-size:0.9em;font-weight:700;color:#1a1a1a;margin-bottom:8px;line-height:1.3;">{f['title']}</div>
      <div style="font-size:0.85em;font-weight:400;line-height:1.6;color:#6b7177;">{f['text']}</div>
    </div>"""
    return f"""<div style="margin-bottom:48px;padding-top:32px;border-top:1px solid #e8e8e8;">
  <p style="font-family:'figmaMono','SF Mono',monospace;font-size:0.7em;font-weight:400;letter-spacing:0.54px;text-transform:uppercase;color:rgba(0,0,0,0.45);margin:0 0 8px 0;">Trivia</p>
  <h2 id="fun-facts" style="font-size:1.5em;font-weight:700;color:#1a1a1a;margin:0 0 24px 0;line-height:1.2;">Fun Facts</h2>
  <div style="display:grid;grid-template-columns:repeat(2,1fr);gap:12px;">{cards}
  </div>
</div>"""


def build_faq(faqs):
    if not faqs:
        return ""
    items = ""
    for i, item in enumerate(faqs):
        border_bottom = ";border-bottom:1px solid #e8e8e8" if i == len(faqs) - 1 else ""
        items += f"""
  <details style="border-top:1px solid #e8e8e8{border_bottom};">
    <summary style="padding:18px 0;font-size:0.95em;font-weight:700;color:#1a1a1a;display:flex;justify-content:space-between;align-items:center;list-style:none;user-select:none;">
      {item['q']}
      <span class="faq-icon" style="font-size:1.4em;font-weight:300;color:#6b7177;flex-shrink:0;margin-left:16px;line-height:1;">+</span>
    </summary>
    <p style="font-size:0.9em;font-weight:400;line-height:1.7;color:#6b7177;margin:0 0 18px 0;padding-right:24px;">{item['a']}</p>
  </details>"""
    # FAQPage JSON-LD schema for AI agents / voice search / Google understanding.
    schema = {
        "@context": "https://schema.org",
        "@type": "FAQPage",
        "mainEntity": [
            {
                "@type": "Question",
                "name": item["q"],
                "acceptedAnswer": {"@type": "Answer", "text": item["a"]},
            }
            for item in faqs
        ],
    }
    schema_json = json.dumps(schema, ensure_ascii=False).replace("</", "<\\/")
    return f"""<style>
.wooffy-faq summary::-webkit-details-marker{{display:none;}}
.wooffy-faq summary{{cursor:pointer;outline:none;}}
.wooffy-faq details[open] .faq-icon{{transform:rotate(45deg);}}
.wooffy-faq .faq-icon{{transition:transform 0.22s ease;display:inline-block;}}
</style>
<script type="application/ld+json">{schema_json}</script>
<div class="wooffy-faq" style="margin-bottom:48px;padding-top:32px;border-top:1px solid #e8e8e8;">
  <p style="font-family:'figmaMono','SF Mono',monospace;font-size:0.7em;font-weight:400;letter-spacing:0.54px;text-transform:uppercase;color:rgba(0,0,0,0.45);margin:0 0 8px 0;">FAQ</p>
  <h2 id="faq" style="font-size:1.5em;font-weight:700;color:#1a1a1a;margin:0 0 4px 0;line-height:1.2;">Frequently Asked Questions</h2>
{items}
</div>"""


def build_related(breeds):
    items = ""
    for b in breeds:
        items += f'    <li><a href="{BLOG_BASE}/{b["slug"]}" style="color:#1a1a1a;font-weight:700;text-decoration:underline;text-underline-offset:3px;">{b["name"]}</a> — {b["note"]}</li>\n'
    return f"""<div style="margin-bottom:48px;padding-top:32px;border-top:1px solid #e8e8e8;">
  <p style="font-family:'figmaMono','SF Mono',monospace;font-size:0.7em;font-weight:400;letter-spacing:0.54px;text-transform:uppercase;color:rgba(0,0,0,0.45);margin:0 0 8px 0;">Explore More</p>
  <h2 style="font-size:1.5em;font-weight:700;color:#1a1a1a;margin:0 0 16px 0;line-height:1.2;">Similar Breeds</h2>
  <ul style="font-size:0.95em;font-weight:400;line-height:1.9;color:#6b7177;padding-left:20px;margin:0;">
{items}  </ul>
</div>"""


def build_image(img, extra_style=""):
    cap = img.get("caption", "")
    cta = img.get("cta", "")
    cta_html = f'<span style="color:#1a1a1a;font-weight:700;text-decoration:underline;text-underline-offset:3px;{"margin-left:8px;" if cap else ""}">{cta}</span>' if cta else ""
    caption = f'<figcaption style="font-size:0.75em;color:#6b7177;margin-top:8px;letter-spacing:0.02em;">{cap}{cta_html}</figcaption>' if (cap or cta) else ""
    figure = f"""<figure style="margin:0 0 8px 0;{extra_style}">
  <img src="{img['url']}" alt="{img['alt']}" loading="lazy"
    style="width:100%;height:auto;border-radius:8px;display:block;box-shadow:rgba(0,0,0,0.07) 0px 8px 24px;">
  {caption}
</figure>"""
    link = img.get("link")
    return f'<a href="{link}" style="text-decoration:none;color:inherit;display:block;">{figure}</a>' if link else figure


def build_quiz_cta(cta):
    illustration = """<svg width="96" height="72" viewBox="0 0 96 72" fill="none" xmlns="http://www.w3.org/2000/svg">
  <circle cx="30" cy="42" r="26" fill="#4D49FC" opacity="0.92"/>
  <circle cx="56" cy="42" r="26" fill="#E4FF97" opacity="0.85"/>
  <circle cx="43" cy="22" r="21" fill="#FF7262" opacity="0.80"/>
  <circle cx="30" cy="42" r="26" fill="none" stroke="rgba(0,0,0,0.08)" stroke-width="1.5"/>
  <circle cx="56" cy="42" r="26" fill="none" stroke="rgba(0,0,0,0.08)" stroke-width="1.5"/>
  <circle cx="43" cy="22" r="21" fill="none" stroke="rgba(0,0,0,0.08)" stroke-width="1.5"/>
</svg>"""
    return f"""<div style="position:relative;overflow:hidden;background:#ffffff;border-radius:12px;border:1px solid rgba(0,0,0,0.08);padding:48px 32px 44px;margin:0 0 48px 0;text-align:center;">
  <div style="position:absolute;width:280px;height:280px;border-radius:50%;background:#4D49FC;filter:blur(72px);opacity:0.22;top:-100px;right:-80px;animation:wooffy-drift1 10s ease-in-out infinite;pointer-events:none;"></div>
  <div style="position:absolute;width:240px;height:240px;border-radius:50%;background:#E4FF97;filter:blur(64px);opacity:0.55;bottom:-80px;left:-60px;animation:wooffy-drift2 13s ease-in-out infinite;pointer-events:none;"></div>
  <div style="position:absolute;width:200px;height:200px;border-radius:50%;background:#FF7262;filter:blur(60px);opacity:0.28;top:30%;left:35%;animation:wooffy-drift3 16s ease-in-out infinite;pointer-events:none;"></div>
  <div style="position:relative;z-index:1;">
    <div style="margin:0 0 22px 0;">{illustration}</div>
    <p style="font-family:'figmaMono','SF Mono',monospace;font-size:0.7em;font-weight:400;letter-spacing:0.54px;text-transform:uppercase;color:rgba(0,0,0,0.45);margin:0 0 12px 0;">Find Your Match</p>
    <h3 style="font-size:1.9em;font-weight:700;color:#000000;margin:0 0 12px 0;line-height:1.15;letter-spacing:-0.44px;">{cta['heading']}</h3>
    <p style="font-size:0.92em;font-weight:400;line-height:1.65;color:rgba(0,0,0,0.58);margin:0 0 28px 0;max-width:460px;margin-left:auto;margin-right:auto;letter-spacing:-0.14px;">{cta['body']}</p>
    <a href="{cta['button_url']}" style="display:inline-block;background:#000000;color:#ffffff;text-decoration:none;padding:12px 28px;border-radius:50px;font-weight:500;font-size:0.88em;letter-spacing:-0.14px;">{cta['button_text']}</a>
  </div>
</div>"""


def build_email_capture(variant):
    if variant == "checklist":
        checklist_svg = """<svg width="96" height="72" viewBox="0 0 96 72" fill="none" xmlns="http://www.w3.org/2000/svg">
  <circle cx="16" cy="18" r="8" fill="#000000"/>
  <path d="M12 18 L15 22 L21 14" stroke="white" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"/>
  <rect x="30" y="14" width="52" height="8" rx="4" fill="#000000" opacity="0.12"/>
  <rect x="30" y="14" width="38" height="8" rx="4" fill="#000000" opacity="0.55"/>
  <circle cx="16" cy="38" r="8" fill="#000000"/>
  <path d="M12 38 L15 42 L21 34" stroke="white" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"/>
  <rect x="30" y="34" width="52" height="8" rx="4" fill="#000000" opacity="0.12"/>
  <rect x="30" y="34" width="46" height="8" rx="4" fill="#000000" opacity="0.55"/>
  <circle cx="16" cy="58" r="8" fill="none" stroke="#000000" stroke-width="2" opacity="0.2"/>
  <rect x="30" y="54" width="52" height="8" rx="4" fill="#000000" opacity="0.08"/>
  <rect x="30" y="54" width="28" height="8" rx="4" fill="#000000" opacity="0.2"/>
</svg>"""
        return f"""<div class="wooffy-upgrade" style="position:relative;overflow:hidden;background:#ffffff;border-radius:12px;border:1px solid rgba(0,0,0,0.08);padding:40px 32px;margin:0 0 48px 0;text-align:center;">
  <div style="position:absolute;width:240px;height:240px;border-radius:50%;background:#E4FF97;filter:blur(68px);opacity:0.65;top:-80px;left:-60px;animation:wooffy-drift2 12s ease-in-out infinite;pointer-events:none;"></div>
  <div style="position:absolute;width:200px;height:200px;border-radius:50%;background:#7DD3FC;filter:blur(60px);opacity:0.40;bottom:-60px;right:-40px;animation:wooffy-drift1 15s ease-in-out infinite;pointer-events:none;"></div>
  <div style="position:absolute;width:180px;height:180px;border-radius:50%;background:#FDE68A;filter:blur(56px);opacity:0.45;top:40%;left:50%;animation:wooffy-drift3 18s ease-in-out infinite;pointer-events:none;"></div>
  <div style="position:relative;z-index:1;">
    <div style="margin:0 0 20px 0;">{checklist_svg}</div>
    <p style="font-family:'figmaMono','SF Mono',monospace;font-size:0.7em;font-weight:400;letter-spacing:0.54px;text-transform:uppercase;color:rgba(0,0,0,0.45);margin:0 0 12px 0;">Free Download</p>
    <h3 style="font-size:1.75em;font-weight:700;color:#000000;margin:0 0 10px 0;line-height:1.25;letter-spacing:-0.44px;">The First-Time Golden Owner Checklist</h3>
    <p style="font-size:0.9em;line-height:1.65;color:rgba(0,0,0,0.58);margin:0 0 24px 0;max-width:400px;margin-left:auto;margin-right:auto;letter-spacing:-0.14px;">47 things to have ready before your puppy comes home—including the stuff nobody tells you about.</p>
    <a href="/pages/coming-soon" style="display:inline-block;background:#000000;color:#ffffff;text-decoration:none;padding:12px 28px;border-radius:50px;font-weight:500;font-size:0.85em;letter-spacing:-0.14px;">Get the checklist</a>
  </div>
</div>"""


def generate_html(breed):
    m   = breed["meta"]
    s   = breed["sections"]
    imgs = breed.get("images", {})
    hero_img = imgs.get("hero", {})
    secondary = {img["placement"]: img for img in imgs.get("secondary", [])}

    name  = m["name"]
    group = m["group"]
    size_cat = m["size_category"]

    hero_html = ""
    hero_figure = ""

    # ── main content ─────────────────────────────────────────────────────────
    content = ""

    # intro (no border-top on first section)
    intro = s.get("intro", {})
    stats_html = f'<div style="margin-top:24px;">{build_stats_grid(breed["stats"])}</div>' if breed.get("stats") else ""
    content += f"""<div style="margin-bottom:48px;">
  <p style="font-family:'figmaMono','SF Mono',monospace;font-size:0.7em;font-weight:400;letter-spacing:0.54px;text-transform:uppercase;color:rgba(0,0,0,0.45);margin:0 0 8px 0;">{intro.get('label','')}</p>
  <h2 id="intro" style="font-size:1.5em;font-weight:700;color:#1a1a1a;margin:0 0 16px 0;line-height:1.2;">{intro.get('heading','')}</h2>
  {intro.get('html','')}
  {stats_html}
</div>"""

    if s.get("appearance"):
        content += build_section("appearance", s["appearance"], name)

    # secondary image before temperament
    if "before_temperament" in secondary:
        content += build_image(secondary["before_temperament"])

    if s.get("temperament"):
        content += build_section("temperament", s["temperament"], name)

    if s.get("mikes_take"):
        content += build_section("mikes_take", s["mikes_take"], name)

    # secondary image before care
    if "before_care" in secondary:
        content += build_image(secondary["before_care"])

    if s.get("care"):
        content += build_section("care",    s["care"],    name)
    if s.get("health"):
        content += build_section("health",  s["health"],  name)
    if s.get("special"):
        content += build_section("special", s["special"], name)
    if s.get("cost"):
        content += build_section("cost",    s["cost"],    name)
    if s.get("right_for_you"):
        content += build_right_for_you(s["right_for_you"])

    # secondary image before finding
    if "before_finding" in secondary:
        content += build_image(secondary["before_finding"])

    if s.get("finding"):
        content += build_section("finding", s["finding"], name)
    content += build_faq(breed.get("faq", []))
    if breed.get("related_breeds"):
        content += build_related(breed["related_breeds"])
    content = add_link_targets(content)
    # ── assemble full page ────────────────────────────────────────────────────
    toc = build_toc(s, breed)

    scripts = """<script>
(function(){
  // ── Keyframe injection (Shopify strips @keyframes from <style> tags) ──────
  var ks = document.createElement('style');
  ks.textContent = [
    '@keyframes wooffy-drift1{0%,100%{transform:translate(0,0) scale(1)}33%{transform:translate(22px,-16px) scale(1.08)}66%{transform:translate(-12px,18px) scale(0.94)}}',
    '@keyframes wooffy-drift2{0%,100%{transform:translate(0,0) scale(1)}40%{transform:translate(-20px,14px) scale(1.12)}70%{transform:translate(16px,-10px) scale(0.92)}}',
    '@keyframes wooffy-drift3{0%,100%{transform:translate(0,0) scale(1)}50%{transform:translate(14px,22px) scale(1.07)}}'
  ].join('');
  document.head.appendChild(ks);

  // ── Header height detection ───────────────────────────────────────────────
  var _hh = 64, _gap = 24;
  function getHeaderHeight() {
    var candidates = ['sticky-header', 'header-wrapper', '.header-wrapper', 'header', '.header', '#header', '[role="banner"]', '.site-header'];
    for (var i = 0; i < candidates.length; i++) {
      var el = document.querySelector(candidates[i]);
      if (el) { var h = Math.round(el.getBoundingClientRect().height); if (h > 10) return h; }
    }
    return 64;
  }
  function calibrateToc() {
    _hh = getHeaderHeight();
    var toc = document.querySelector('.wooffy-toc');
    if (toc) {
      toc.style.top = (_hh + _gap) + 'px';
      toc.style.maxHeight = 'calc(100vh - ' + (_hh + _gap + 16) + 'px)';
    }
    document.querySelectorAll('.wooffy-main h2, .wooffy-main h3').forEach(function(el) {
      el.style.scrollMarginTop = (_hh + _gap) + 'px';
    });
  }
  // Run immediately, then again after Shopify's JS header finishes rendering
  calibrateToc();
  setTimeout(calibrateToc, 300);
  setTimeout(calibrateToc, 800);
  window.addEventListener('resize', calibrateToc, {passive:true});

  // ── TOC scroll sync using getBoundingClientRect (viewport-relative) ───────
  // offsetTop is unreliable (relative to nearest positioned ancestor).
  // getBoundingClientRect().top directly gives viewport position — always correct.
  var links = document.querySelectorAll('.wooffy-toc a[href^="#"]');
  var ids = Array.from(links).map(function(l){ return l.getAttribute('href').slice(1); });
  var sections = ids.map(function(id){ return document.getElementById(id); }).filter(Boolean);
  function update() {
    // A section is "active" once its top edge has scrolled past the header.
    // threshold = header height + gap + tiny buffer (4px).
    var threshold = _hh + _gap + 4;
    var active = sections[0];
    sections.forEach(function(s) {
      if (s.getBoundingClientRect().top <= threshold) active = s;
    });
    links.forEach(function(l) {
      l.classList.toggle('toc-active', l.getAttribute('href') === '#' + active.id);
    });
  }
  window.addEventListener('scroll', update, {passive:true});
  update();
})();
</script>"""

    return f"""<style>
/* ── Figma-inspired design system ── */
html {{ scroll-behavior: smooth; }}
.wooffy-article {{
  font-family: 'figmaSans', 'SF Pro Display', system-ui, 'Helvetica Neue', Arial, sans-serif;
  color: #000000;
  line-height: 1.7;
}}
.wooffy-article p,
.wooffy-article li,
.wooffy-article td,
.wooffy-article th {{ line-height: 1.7 !important; color: rgba(0,0,0,0.65); letter-spacing: -0.14px; }}
.wooffy-article h1 {{ line-height: 1.05 !important; letter-spacing: -0.96px; color: #000000; }}
.wooffy-article h2 {{ line-height: 1.2 !important; letter-spacing: -0.44px; color: #000000; }}
.wooffy-article h3 {{ line-height: 1.3 !important; letter-spacing: -0.26px; color: #000000; }}
.wooffy-article strong {{ color: #000000; }}
.wooffy-article a {{ color: #000000; }}

.wooffy-layout {{
  display: grid;
  grid-template-columns: 190px 1fr;
  gap: 56px;
  align-items: start;
}}
.wooffy-toc {{
  position: sticky;
  top: 80px;
  max-height: calc(100vh - 100px);
  overflow-y: auto;
}}
.wooffy-main h2,
.wooffy-main h3 {{
  scroll-margin-top: 88px;
}}
.wooffy-toc ol {{
  margin: 0;
  padding: 0;
  list-style: none;
}}
.wooffy-toc ol li {{
  margin-bottom: 0;
  line-height: 1.4 !important;
}}
.wooffy-toc ol li a {{
  display: block;
  position: relative;
  font-size: 0.82em;
  color: #c0c0c0;
  text-decoration: none;
  padding: 6px 0;
  padding-left: 12px;
  border-left: 2px solid transparent;
  transition: color 0.2s, border-color 0.2s;
}}
.wooffy-toc ol li a:hover {{ color: #444444; }}
.wooffy-toc ol li a.toc-active {{
  color: #000000;
  font-weight: 600;
  border-left-color: #000000;
}}

.wooffy-article table {{
  border-radius: 10px;
  border-collapse: separate !important;
  border-spacing: 0;
  overflow: hidden;
  border: 1px solid rgba(0,0,0,0.09);
  width: 100%;
}}
.wooffy-article table th:first-child {{ border-radius: 10px 0 0 0; }}
.wooffy-article table th:last-child  {{ border-radius: 0 10px 0 0; }}
.wooffy-article table tr:last-child td:first-child {{ border-radius: 0 0 0 10px; }}
.wooffy-article table tr:last-child td:last-child  {{ border-radius: 0 0 10px 0; }}

@media (max-width: 900px) {{
  .wooffy-layout {{ display: block; }}
  .wooffy-toc {{ display: none; }}
}}

</style>
<div class="wooffy-article" style="padding:40px 0;margin:0;">

{hero_html}

{hero_figure}

<!-- LAYOUT: STICKY SIDEBAR + MAIN -->
<div class="wooffy-layout">

{toc}

<div class="wooffy-main">
{content}
</div><!-- /wooffy-main -->
</div><!-- /wooffy-layout -->

{scripts}
</div><!-- /wooffy-article -->"""


def save_html(slug, html):
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    out_path = os.path.join(OUTPUT_DIR, f"{slug}.html")
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"  Saved -> {out_path}")
    return out_path


# ── Published log helpers ─────────────────────────────────────────────────────

def load_log():
    if os.path.exists(LOG_PATH):
        with open(LOG_PATH, encoding="utf-8") as f:
            return json.load(f)
    return {}

def save_log(log):
    with open(LOG_PATH, "w", encoding="utf-8") as f:
        json.dump(log, f, indent=2)

# ── Shopify API helpers ───────────────────────────────────────────────────────

def _shopify_request(path, method="GET", payload=None):
    """Low-level Shopify Admin API call. Raises on HTTP errors."""
    url = f"https://{SHOPIFY_STORE}/admin/api/2024-01{path}"
    data = json.dumps(payload).encode("utf-8") if payload else None
    req = urllib.request.Request(url, data=data, method=method)
    req.add_header("X-Shopify-Access-Token", SHOPIFY_TOKEN)
    req.add_header("Content-Type", "application/json")
    try:
        with urllib.request.urlopen(req) as resp:
            return json.loads(resp.read())
    except urllib.error.HTTPError as e:
        body = e.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"Shopify API {e.code}: {body}") from e


def update_articles_md():
    """Regenerate ARTICLES.md from all breed-data/*.json files."""
    files = sorted(glob.glob(os.path.join(BREED_DATA_DIR, "*.json")))
    articles = []
    for f in files:
        with open(f, encoding="utf-8") as fp:
            data = json.load(fp)
        m = data.get("meta", {})
        articles.append({
            "name":          m.get("name", ""),
            "slug":          m.get("slug", ""),
            "group":         m.get("group", ""),
            "size_category": m.get("size_category", ""),
            "published_at":  m.get("published_at", None),
        })

    # ── classify ──────────────────────────────────────────────────────────────
    SUPPORT_SUFFIXES = ("-grooming-guide", "-first-year-costs", "-puppy-checklist", "-outdoor-space")
    ROUNDUP_SIZES    = ("Roundup",)

    roundups  = [a for a in articles if a["size_category"] in ROUNDUP_SIZES]
    guides    = [a for a in articles if not any(a["slug"].endswith(s) for s in SUPPORT_SUFFIXES)
                                     and a["size_category"] not in ROUNDUP_SIZES]
    supports  = [a for a in articles if any(a["slug"].endswith(s) for s in SUPPORT_SUFFIXES)]

    today = datetime.utcnow().strftime("%Y-%m-%d")

    now_utc = datetime.utcnow()

    def status_str(a):
        if a["published_at"]:
            try:
                dt = datetime.strptime(a["published_at"], "%Y-%m-%dT%H:%M:%SZ")
                if dt > now_utc:
                    # still in the future — show scheduled time in Beijing
                    bjt_hour = (dt.hour + 8) % 24
                    bjt_day  = dt.strftime("%Y-%m-%d")
                    return f"🕐 Scheduled {bjt_day} {bjt_hour:02d}:00 BJT"
            except Exception:
                pass
        return "✅ Live"

    def article_row(a):
        url = f"https://thewooffy.com{BLOG_BASE}/{a['slug']}"
        return f"| {a['name']} | {status_str(a)} | [{BLOG_BASE}/{a['slug']}]({url}) |"

    # ── group by breed slug (not AKC group — multiple breeds share the same group) ──
    # guide_order: sorted list of guide slugs (alphabetical)
    # breed_map:   slug → { guide, supports, akc_group, size }
    guide_order = sorted([a["slug"] for a in guides])
    breed_map   = {}
    for a in guides:
        breed_map[a["slug"]] = {
            "guide":     a,
            "supports":  [],
            "akc_group": a["group"],
            "size":      a["size_category"],
        }
    for a in supports:
        # find which guide this support belongs to by slug prefix match
        best = None
        for guide_slug in breed_map:
            if a["slug"].startswith(guide_slug + "-"):
                if best is None or len(guide_slug) > len(best):
                    best = guide_slug
        if best:
            breed_map[best]["supports"].append(a)

    lines = []
    lines.append("# Wooffy Article Tracker")
    lines.append("")
    lines.append("> Auto-generated by `scripts/generate.py`. Do not edit manually.")
    lines.append(f"> Last updated: {today}")
    lines.append("")
    lines.append("## Summary")
    lines.append("")
    lines.append("| Type | Count |")
    lines.append("|------|-------|")
    lines.append(f"| Breed guides | {len(guides)} |")
    lines.append(f"| Supporting articles | {len(supports)} |")
    lines.append(f"| Roundup articles | {len(roundups)} |")
    lines.append(f"| **Total** | **{len(articles)}** |")
    lines.append("")
    lines.append("---")
    lines.append("")

    if roundups:
        lines.append("## Roundup Articles")
        lines.append("")
        lines.append("| Article | Status | URL |")
        lines.append("|---------|--------|-----|")
        for a in roundups:
            lines.append(article_row(a))
        lines.append("")
        lines.append("---")
        lines.append("")

    lines.append("## Breed Guides")
    lines.append("")
    for slug in guide_order:
        bdata = breed_map[slug]
        guide = bdata["guide"]
        lines.append(f"### {guide['name']} *({bdata['akc_group']} · {bdata['size']})*")
        lines.append("| Article | Status | URL |")
        lines.append("|---------|--------|-----|")
        lines.append(article_row(guide))
        for s in sorted(bdata["supports"], key=lambda x: x["slug"]):
            lines.append(article_row(s))
        lines.append("")

    with open(ARTICLES_PATH, "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")
    print(f"  ARTICLES.md updated ({len(articles)} articles total)")


def update_articles_xlsx():
    """Regenerate ARTICLES.xlsx from all breed-data/*.json files."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  SKIP ARTICLES.xlsx — openpyxl not installed (pip install openpyxl)")
        return

    files = sorted(glob.glob(os.path.join(BREED_DATA_DIR, "*.json")))
    articles = []
    for f in files:
        with open(f, encoding="utf-8") as fp:
            data = json.load(fp)
        m = data.get("meta", {})
        slug   = m.get("slug", "")
        pub_at = m.get("published_at", None)
        status = "Live"
        if pub_at:
            try:
                dt = datetime.strptime(pub_at, "%Y-%m-%dT%H:%M:%SZ")
                if dt > datetime.utcnow():
                    bjt_hour = (dt.hour + 8) % 24
                    status   = f"Scheduled {dt.strftime('%Y-%m-%d')} {bjt_hour:02d}:00 BJT"
            except Exception:
                pass

        size = m.get("size_category", "")
        SUPPORT_SUFFIXES = ("-grooming-guide", "-first-year-costs", "-puppy-checklist", "-outdoor-space")
        if size in ("Roundup",):
            article_type = "Roundup"
        elif any(slug.endswith(s) for s in SUPPORT_SUFFIXES):
            article_type = "Supporting"
        else:
            article_type = "Breed Guide"

        articles.append({
            "name":   m.get("name", ""),
            "slug":   slug,
            "type":   article_type,
            "group":  m.get("group", ""),
            "size":   size,
            "status": status,
            "url":    f"https://{SHOPIFY_STORE}{BLOG_BASE}/{slug}",
        })

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Articles"

    thin        = Side(style="thin", color="D0D0D0")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1A1A1A")
    header_align = Alignment(horizontal="center", vertical="center")

    headers    = ["#", "Article Name", "Type", "AKC Group", "Size", "Status", "URL"]
    col_widths = [4, 52, 14, 22, 16, 32, 70]

    ws.row_dimensions[1].height = 28
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        cell.border    = border
        ws.column_dimensions[get_column_letter(col)].width = w

    for i, a in enumerate(articles, 1):
        row = i + 1
        ws.row_dimensions[row].height = 20

        row_color    = "F0F4FF" if a["type"] == "Breed Guide" else ("FFF8E1" if a["type"] == "Roundup" else "FFFFFF")
        status_color = "E8F5E9" if a["status"] == "Live" else "FFF3E0"

        values = [i, a["name"], a["type"], a["group"], a["size"], a["status"], a["url"]]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border    = border
            cell.alignment = Alignment(vertical="center")
            cell.font      = Font(name="Calibri", size=10)
            cell.fill      = PatternFill("solid", fgColor=status_color if col == 6 else row_color)

        url_cell           = ws.cell(row=row, column=7)
        url_cell.hyperlink = a["url"]
        url_cell.font      = Font(name="Calibri", size=10, color="0563C1", underline="single")

    ws.freeze_panes         = "A2"
    ws.auto_filter.ref      = f"A1:G{len(articles)+1}"

    wb.save(XLSX_PATH)
    print(f"  ARTICLES.xlsx updated ({len(articles)} rows)")


def get_blog_id():
    """List all blogs in the store to find SHOPIFY_BLOG_ID."""
    data = _shopify_request("/blogs.json")
    print("\nBlogs in your store:")
    for blog in data["blogs"]:
        print(f"  id={blog['id']}  title={blog['title']}  handle={blog['handle']}")
    print("\nPaste the correct id into .env as SHOPIFY_BLOG_ID=<id>")


def find_article_by_handle(handle, blog_id=None):
    """Return article_id if handle already exists in the blog, else None."""
    blog_id = blog_id or SHOPIFY_BLOG_ID
    data = _shopify_request(
        f"/blogs/{blog_id}/articles.json?handle={handle}&fields=id,handle"
    )
    for article in data.get("articles", []):
        if article["handle"] == handle:
            return article["id"]
    return None


def _compute_title_tag(slug, name):
    """SEO title_tag template covering breeds (main+supporting), roundups, and comparisons."""
    roundup_prefixes = ("best-", "most-", "easiest-", "quietest-", "longest-", "rarest-", "low-", "dog-breeds-")
    if any(slug.startswith(p) for p in roundup_prefixes):
        return f"{name}: Top Picks & Care Guide"
    if "-vs-" in slug:
        return f"{name}: Which Is Right for You?"
    suffix_templates = {
        "-grooming-guide":    "{breed} Grooming Guide: Tools, Tips & Schedule",
        "-first-year-costs":  "{breed} Cost: First-Year & Annual Budget",
        "-puppy-checklist":   "{breed} Puppy Checklist: First 30 Days Essentials",
    }
    for suf, tmpl in suffix_templates.items():
        if slug.endswith(suf):
            base_slug = slug[: -len(suf)]
            breed_display = base_slug.replace("-", " ").title()
            return tmpl.format(breed=breed_display)
    return f"{name} Breed Guide: Cost, Care & Temperament"


def publish_to_shopify(breed, html, force_update=False):
    m = breed["meta"]
    handle = m["shopify_handle"]
    blog_handle, blog_id = _resolve_blog(breed)
    log = load_log()

    # ── duplicate check ───────────────────────────────────────────────────────
    existing_id = find_article_by_handle(handle, blog_id)

    print(f"  excerpt preview: {m.get('excerpt','')[:60] or '(empty)'}")

    if existing_id and not force_update:
        print(f"  SKIP (already published, id={existing_id}). Use --update to overwrite.")
        return existing_id

    hero = breed.get("images", {}).get("hero", {})
    # Prefer an explicit title_tag stored in the JSON (e.g. imported nutrition
    # articles already carry the real SERP title metafield). Fall back to the
    # template-based computation for breeds/grooming/cost/checklist/roundups.
    title_tag = m.get("title_tag") or _compute_title_tag(handle, m["name"])
    article_body = {
        "title":      m["name"],
        "body_html":  html,
        "handle":     handle,
        "author":     m.get("author", "Kailun Zhang"),
        "published":  m.get("published", True),
        "tags":       ", ".join(m.get("tags", [])),
        "excerpt":    m.get("excerpt", ""),
        **({"published_at": m["published_at"]} if m.get("published_at") else {}),
        **({"image": {"src": hero["url"], "alt": hero.get("alt", m["name"])}} if hero.get("url") else {}),
        "metafields": [
            mf for mf in [
                ({
                    "key":       "description_tag",
                    "value":     m.get("meta_description", ""),
                    "type":      "single_line_text_field",
                    "namespace": "global"
                } if m.get("meta_description") else None),
                ({
                    "key":       "title_tag",
                    "value":     title_tag,
                    "type":      "single_line_text_field",
                    "namespace": "global"
                } if title_tag else None),
            ] if mf is not None
        ]
    }

    if existing_id:
        # UPDATE existing article
        result = _shopify_request(
            f"/blogs/{blog_id}/articles/{existing_id}.json",
            method="PUT",
            payload={"article": article_body}
        )
        article_id = result["article"]["id"]
        action = "Updated"
    else:
        # CREATE new article
        result = _shopify_request(
            f"/blogs/{blog_id}/articles.json",
            method="POST",
            payload={"article": article_body}
        )
        article_id = result["article"]["id"]
        action = "Published"

    url = f"https://{SHOPIFY_STORE}/blogs/{blog_handle}/{handle}"
    print(f"  {action} -> {url}  (id={article_id})")

    # ── write to log ──────────────────────────────────────────────────────────
    log[m["slug"]] = {
        "article_id": article_id,
        "handle":     handle,
        "url":        url,
        "action":     action.lower(),
        "timestamp":  datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ"),
    }
    save_log(log)
    return article_id


# ── CLI entry point ────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import time

    args = sys.argv[1:]

    if not args or "--help" in args:
        print(__doc__)
        sys.exit(0)

    if "--get-blog-id" in args:
        get_blog_id()
        sys.exit(0)

    if "--update-articles-md" in args:
        update_articles_md()
        update_articles_xlsx()
        sys.exit(0)

    do_publish = "--publish" in args
    do_update  = "--update"  in args
    slugs      = [a for a in args if not a.startswith("--")]

    if "--all" in args:
        slugs = [os.path.splitext(os.path.basename(f))[0]
                 for f in sorted(glob.glob(os.path.join(BREED_DATA_DIR, "*.json")))]

    if not slugs:
        print("No breed slugs specified. Use --all or pass a slug name.")
        sys.exit(1)

    # ── confirm before publishing ─────────────────────────────────────────────
    if do_publish or do_update:
        if not SHOPIFY_TOKEN or SHOPIFY_TOKEN == "shpat_xxxx":
            print("ERROR: SHOPIFY_TOKEN not set in .env")
            sys.exit(1)
        if SHOPIFY_BLOG_ID == 0:
            print("ERROR: SHOPIFY_BLOG_ID not set in .env  (run --get-blog-id first)")
            sys.exit(1)
        action_label = "UPDATE" if do_update else "PUBLISH"
        print(f"\nStore : {SHOPIFY_STORE}")
        print(f"Blog  : {SHOPIFY_BLOG_ID}")
        print(f"Action: {action_label}")
        print(f"Breeds: {', '.join(slugs)}")
        confirm = input("\nType YES to continue: ").strip()
        if confirm != "YES":
            print("Aborted.")
            sys.exit(0)

    ok, skipped, failed = 0, 0, 0

    for slug in slugs:
        print(f"\n>> {slug}")
        try:
            breed = load_breed(slug)
            if breed.get("body_html") is not None:
                # Passthrough mode (imported nutrition / non-wooffy templates):
                # the JSON carries the article HTML verbatim. Do NOT render the
                # wooffy section template — it would overwrite the real body.
                html = breed["body_html"]
                print(f"  passthrough mode  (body_html, {len(html)} bytes)")
            else:
                html = generate_html(breed)
            save_html(slug, html)
            if do_publish or do_update:
                publish_to_shopify(breed, html, force_update=do_update)
                time.sleep(0.6)
            ok += 1
        except Exception as e:
            print(f"  ERROR: {e}")
            failed += 1

    print(f"\nDone. ok={ok}  skipped={skipped}  failed={failed}")

    if (do_publish or do_update) and ok > 0:
        update_articles_md()
        update_articles_xlsx()
