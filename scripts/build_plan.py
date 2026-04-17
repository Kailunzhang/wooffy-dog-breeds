"""
Build BREED_PLAN.xlsx — full roadmap of all planned breeds and articles.
Usage: python scripts/build_plan.py
"""

import json, os, glob
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT_DIR       = os.path.join(os.path.dirname(__file__), "..")
BREED_DATA_DIR = os.path.join(ROOT_DIR, "breed-data")
BLOG_BASE      = "https://thewooffy.com/blogs/dog-breeds"
OUT_PATH       = os.path.join(ROOT_DIR, "BREED_PLAN.xlsx")

# ── Master breed list ─────────────────────────────────────────────────────────
# Format: (Breed Name, AKC Group, Size, Priority)
# Priority: High = popular/high search volume, Medium = moderate, Low = niche
BREEDS = [
    # Sporting Group
    ("Brittany",                        "Sporting Group", "Medium Breed", "Medium"),
    ("Pointer",                         "Sporting Group", "Large Breed",  "Low"),
    ("German Shorthaired Pointer",      "Sporting Group", "Large Breed",  "High"),
    ("German Wirehaired Pointer",       "Sporting Group", "Large Breed",  "Medium"),
    ("Chesapeake Bay Retriever",        "Sporting Group", "Large Breed",  "Medium"),
    ("Flat-Coated Retriever",           "Sporting Group", "Large Breed",  "Medium"),
    ("Golden Retriever",                "Sporting Group", "Large Breed",  "High"),
    ("Labrador Retriever",              "Sporting Group", "Large Breed",  "High"),
    ("Nova Scotia Duck Tolling Retriever","Sporting Group","Medium Breed","Medium"),
    ("English Setter",                  "Sporting Group", "Large Breed",  "Low"),
    ("Gordon Setter",                   "Sporting Group", "Large Breed",  "Low"),
    ("Irish Setter",                    "Sporting Group", "Large Breed",  "Medium"),
    ("American Water Spaniel",          "Sporting Group", "Medium Breed", "Low"),
    ("Clumber Spaniel",                 "Sporting Group", "Medium Breed", "Low"),
    ("Cocker Spaniel",                  "Sporting Group", "Medium Breed", "High"),
    ("English Cocker Spaniel",          "Sporting Group", "Medium Breed", "Medium"),
    ("English Springer Spaniel",        "Sporting Group", "Medium Breed", "High"),
    ("Field Spaniel",                   "Sporting Group", "Medium Breed", "Low"),
    ("Irish Water Spaniel",             "Sporting Group", "Large Breed",  "Low"),
    ("Welsh Springer Spaniel",          "Sporting Group", "Medium Breed", "Low"),
    ("Vizsla",                          "Sporting Group", "Medium Breed", "High"),
    ("Weimaraner",                      "Sporting Group", "Large Breed",  "High"),
    ("Wirehaired Pointing Griffon",     "Sporting Group", "Medium Breed", "Low"),
    ("Lagotto Romagnolo",               "Sporting Group", "Medium Breed", "Medium"),
    ("Spinone Italiano",                "Sporting Group", "Large Breed",  "Low"),

    # Hound Group
    ("Afghan Hound",                    "Hound Group",    "Large Breed",  "Medium"),
    ("American Foxhound",               "Hound Group",    "Large Breed",  "Low"),
    ("Basenji",                         "Hound Group",    "Small Breed",  "Medium"),
    ("Basset Hound",                    "Hound Group",    "Medium Breed", "High"),
    ("Beagle",                          "Hound Group",    "Small Breed",  "High"),
    ("Black and Tan Coonhound",         "Hound Group",    "Large Breed",  "Low"),
    ("Bloodhound",                      "Hound Group",    "Large Breed",  "Medium"),
    ("Bluetick Coonhound",              "Hound Group",    "Large Breed",  "Low"),
    ("Borzoi",                          "Hound Group",    "Large Breed",  "Medium"),
    ("Dachshund",                       "Hound Group",    "Small Breed",  "High"),
    ("Greyhound",                       "Hound Group",    "Large Breed",  "High"),
    ("Harrier",                         "Hound Group",    "Medium Breed", "Low"),
    ("Ibizan Hound",                    "Hound Group",    "Large Breed",  "Low"),
    ("Irish Wolfhound",                 "Hound Group",    "Giant Breed",  "Medium"),
    ("Norwegian Elkhound",              "Hound Group",    "Medium Breed", "Low"),
    ("Otterhound",                      "Hound Group",    "Large Breed",  "Low"),
    ("Pharaoh Hound",                   "Hound Group",    "Medium Breed", "Low"),
    ("Redbone Coonhound",               "Hound Group",    "Large Breed",  "Low"),
    ("Rhodesian Ridgeback",             "Hound Group",    "Large Breed",  "High"),
    ("Saluki",                          "Hound Group",    "Large Breed",  "Medium"),
    ("Scottish Deerhound",              "Hound Group",    "Giant Breed",  "Low"),
    ("Whippet",                         "Hound Group",    "Medium Breed", "High"),

    # Working Group
    ("Akita",                           "Working Group",  "Large Breed",  "High"),
    ("Alaskan Malamute",                "Working Group",  "Large Breed",  "High"),
    ("Anatolian Shepherd Dog",          "Working Group",  "Large Breed",  "Medium"),
    ("Bernese Mountain Dog",            "Working Group",  "Large Breed",  "High"),
    ("Boxer",                           "Working Group",  "Large Breed",  "High"),
    ("Bullmastiff",                     "Working Group",  "Large Breed",  "Medium"),
    ("Cane Corso",                      "Working Group",  "Large Breed",  "High"),
    ("Doberman Pinscher",               "Working Group",  "Large Breed",  "High"),
    ("Dogue de Bordeaux",               "Working Group",  "Large Breed",  "Medium"),
    ("German Pinscher",                 "Working Group",  "Medium Breed", "Low"),
    ("Giant Schnauzer",                 "Working Group",  "Large Breed",  "Medium"),
    ("Great Dane",                      "Working Group",  "Giant Breed",  "High"),
    ("Great Pyrenees",                  "Working Group",  "Large Breed",  "High"),
    ("Greater Swiss Mountain Dog",      "Working Group",  "Large Breed",  "Medium"),
    ("Komondor",                        "Working Group",  "Large Breed",  "Low"),
    ("Leonberger",                      "Working Group",  "Giant Breed",  "Medium"),
    ("Mastiff",                         "Working Group",  "Giant Breed",  "High"),
    ("Neapolitan Mastiff",              "Working Group",  "Giant Breed",  "Medium"),
    ("Newfoundland",                    "Working Group",  "Giant Breed",  "High"),
    ("Portuguese Water Dog",            "Working Group",  "Medium Breed", "High"),
    ("Rottweiler",                      "Working Group",  "Large Breed",  "High"),
    ("Saint Bernard",                   "Working Group",  "Giant Breed",  "High"),
    ("Samoyed",                         "Working Group",  "Large Breed",  "High"),
    ("Siberian Husky",                  "Working Group",  "Medium Breed", "High"),
    ("Standard Schnauzer",              "Working Group",  "Medium Breed", "Medium"),
    ("Tibetan Mastiff",                 "Working Group",  "Giant Breed",  "Medium"),

    # Terrier Group
    ("Airedale Terrier",                "Terrier Group",  "Medium Breed", "Medium"),
    ("American Staffordshire Terrier",  "Terrier Group",  "Medium Breed", "High"),
    ("Australian Terrier",              "Terrier Group",  "Small Breed",  "Low"),
    ("Bedlington Terrier",              "Terrier Group",  "Small Breed",  "Low"),
    ("Border Terrier",                  "Terrier Group",  "Small Breed",  "Medium"),
    ("Bull Terrier",                    "Terrier Group",  "Medium Breed", "High"),
    ("Cairn Terrier",                   "Terrier Group",  "Small Breed",  "Medium"),
    ("Irish Terrier",                   "Terrier Group",  "Medium Breed", "Low"),
    ("Kerry Blue Terrier",              "Terrier Group",  "Medium Breed", "Low"),
    ("Lakeland Terrier",                "Terrier Group",  "Small Breed",  "Low"),
    ("Miniature Bull Terrier",          "Terrier Group",  "Small Breed",  "Medium"),
    ("Miniature Schnauzer",             "Terrier Group",  "Small Breed",  "High"),
    ("Norfolk Terrier",                 "Terrier Group",  "Small Breed",  "Low"),
    ("Norwich Terrier",                 "Terrier Group",  "Small Breed",  "Low"),
    ("Parson Russell Terrier",          "Terrier Group",  "Small Breed",  "Medium"),
    ("Russell Terrier",                 "Terrier Group",  "Small Breed",  "Medium"),
    ("Scottish Terrier",                "Terrier Group",  "Small Breed",  "Medium"),
    ("Skye Terrier",                    "Terrier Group",  "Small Breed",  "Low"),
    ("Soft Coated Wheaten Terrier",     "Terrier Group",  "Medium Breed", "High"),
    ("Staffordshire Bull Terrier",      "Terrier Group",  "Medium Breed", "High"),
    ("Welsh Terrier",                   "Terrier Group",  "Small Breed",  "Low"),
    ("West Highland White Terrier",     "Terrier Group",  "Small Breed",  "High"),
    ("Wire Fox Terrier",                "Terrier Group",  "Small Breed",  "Low"),

    # Toy Group
    ("Affenpinscher",                   "Toy Group",      "Small Breed",  "Low"),
    ("Brussels Griffon",                "Toy Group",      "Small Breed",  "Medium"),
    ("Cavalier King Charles Spaniel",   "Toy Group",      "Small Breed",  "High"),
    ("Chihuahua",                       "Toy Group",      "Small Breed",  "High"),
    ("Chinese Crested",                 "Toy Group",      "Small Breed",  "Medium"),
    ("Havanese",                        "Toy Group",      "Small Breed",  "High"),
    ("Italian Greyhound",               "Toy Group",      "Small Breed",  "High"),
    ("Japanese Chin",                   "Toy Group",      "Small Breed",  "Low"),
    ("Maltese",                         "Toy Group",      "Small Breed",  "High"),
    ("Miniature Pinscher",              "Toy Group",      "Small Breed",  "Medium"),
    ("Papillon",                        "Toy Group",      "Small Breed",  "Medium"),
    ("Pekingese",                       "Toy Group",      "Small Breed",  "Medium"),
    ("Pomeranian",                      "Toy Group",      "Small Breed",  "High"),
    ("Pug",                             "Toy Group",      "Small Breed",  "High"),
    ("Shih Tzu",                        "Toy Group",      "Small Breed",  "High"),
    ("Silky Terrier",                   "Toy Group",      "Small Breed",  "Low"),
    ("Toy Fox Terrier",                 "Toy Group",      "Small Breed",  "Low"),
    ("Yorkshire Terrier",               "Toy Group",      "Small Breed",  "High"),

    # Non-Sporting Group
    ("American Eskimo Dog",             "Non-Sporting Group","Small Breed","Medium"),
    ("Bichon Frise",                    "Non-Sporting Group","Small Breed","High"),
    ("Boston Terrier",                  "Non-Sporting Group","Small Breed","High"),
    ("Bulldog",                         "Non-Sporting Group","Medium Breed","High"),
    ("Chinese Shar-Pei",                "Non-Sporting Group","Medium Breed","Medium"),
    ("Chow Chow",                       "Non-Sporting Group","Large Breed","High"),
    ("Coton de Tulear",                 "Non-Sporting Group","Small Breed","Medium"),
    ("Dalmatian",                       "Non-Sporting Group","Large Breed","High"),
    ("Finnish Spitz",                   "Non-Sporting Group","Medium Breed","Low"),
    ("French Bulldog",                  "Non-Sporting Group","Small Breed","High"),
    ("Keeshond",                        "Non-Sporting Group","Medium Breed","Low"),
    ("Lhasa Apso",                      "Non-Sporting Group","Small Breed","Medium"),
    ("Lowchen",                         "Non-Sporting Group","Small Breed","Low"),
    ("Miniature Poodle",                "Non-Sporting Group","Small Breed","High"),
    ("Standard Poodle",                 "Non-Sporting Group","Large Breed","High"),
    ("Schipperke",                      "Non-Sporting Group","Small Breed","Low"),
    ("Shiba Inu",                       "Non-Sporting Group","Small Breed","High"),
    ("Tibetan Spaniel",                 "Non-Sporting Group","Small Breed","Low"),
    ("Tibetan Terrier",                 "Non-Sporting Group","Medium Breed","Low"),
    ("Xoloitzcuintli",                  "Non-Sporting Group","Small Breed","Low"),

    # Herding Group
    ("Australian Cattle Dog",           "Herding Group",  "Medium Breed", "High"),
    ("Australian Shepherd",             "Herding Group",  "Medium Breed", "High"),
    ("Bearded Collie",                  "Herding Group",  "Medium Breed", "Medium"),
    ("Beauceron",                       "Herding Group",  "Large Breed",  "Low"),
    ("Belgian Malinois",                "Herding Group",  "Large Breed",  "High"),
    ("Belgian Sheepdog",                "Herding Group",  "Large Breed",  "Low"),
    ("Belgian Tervuren",                "Herding Group",  "Large Breed",  "Low"),
    ("Border Collie",                   "Herding Group",  "Medium Breed", "High"),
    ("Bouvier des Flandres",            "Herding Group",  "Large Breed",  "Medium"),
    ("Briard",                          "Herding Group",  "Large Breed",  "Low"),
    ("Cardigan Welsh Corgi",            "Herding Group",  "Small Breed",  "High"),
    ("Collie",                          "Herding Group",  "Large Breed",  "High"),
    ("Entlebucher Mountain Dog",        "Herding Group",  "Medium Breed", "Low"),
    ("Finnish Lapphund",                "Herding Group",  "Medium Breed", "Low"),
    ("German Shepherd Dog",             "Herding Group",  "Large Breed",  "High"),
    ("Icelandic Sheepdog",              "Herding Group",  "Medium Breed", "Low"),
    ("Miniature American Shepherd",     "Herding Group",  "Small Breed",  "High"),
    ("Old English Sheepdog",            "Herding Group",  "Large Breed",  "High"),
    ("Pembroke Welsh Corgi",            "Herding Group",  "Small Breed",  "High"),
    ("Polish Lowland Sheepdog",         "Herding Group",  "Medium Breed", "Low"),
    ("Puli",                            "Herding Group",  "Medium Breed", "Low"),
    ("Shetland Sheepdog",               "Herding Group",  "Small Breed",  "High"),
    ("Spanish Water Dog",               "Herding Group",  "Medium Breed", "Low"),
    ("Swedish Vallhund",                "Herding Group",  "Small Breed",  "Low"),
]


def name_to_slug(name):
    """Convert breed name to URL slug."""
    return name.lower().replace(" ", "-").replace("'", "").replace("(", "").replace(")", "")


def get_existing_slugs():
    """Return set of slugs that already have JSON files."""
    files = glob.glob(os.path.join(BREED_DATA_DIR, "*.json"))
    slugs = set()
    for f in files:
        with open(f, encoding="utf-8") as fp:
            data = json.load(fp)
        m = data.get("meta", {})
        slug = m.get("slug", "")
        pub_at = m.get("published_at", None)
        size = m.get("size_category", "")
        # only count as "done" if it's a main breed guide (not support or roundup)
        SUPPORT_SUFFIXES = ("-grooming-guide", "-first-year-costs", "-puppy-checklist", "-outdoor-space")
        if size not in ("Roundup", "Guide") and not any(slug.endswith(s) for s in SUPPORT_SUFFIXES):
            is_future = False
            if pub_at:
                try:
                    dt = datetime.strptime(pub_at, "%Y-%m-%dT%H:%M:%SZ")
                    is_future = dt > datetime.utcnow()
                except Exception:
                    pass
            slugs.add((slug, "scheduled" if is_future else "live"))
    return slugs


def build_plan():
    existing = get_existing_slugs()
    existing_slugs = {s for s, _ in existing}
    slug_status    = {s: st for s, st in existing}

    wb = openpyxl.Workbook()

    # ── Sheet 1: Breed Overview ───────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Breed Overview"

    thin   = Side(style="thin", color="D8D8D8")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hcell(ws, row, col, value, bg="1A1A1A", fg="FFFFFF", bold=True, size=11, align="center"):
        c = ws.cell(row=row, column=col, value=value)
        c.font      = Font(name="Calibri", bold=bold, color=fg, size=size)
        c.fill      = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=False)
        c.border    = border
        return c

    def dcell(ws, row, col, value, bg="FFFFFF", bold=False, size=10, align="left", color="1A1A1A", hyperlink=None):
        c = ws.cell(row=row, column=col, value=value)
        c.font      = Font(name="Calibri", bold=bold, color=color, size=size, underline="single" if hyperlink else None)
        c.fill      = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=False)
        c.border    = border
        if hyperlink:
            c.hyperlink = hyperlink
        return c

    # colors
    C_HEADER  = "1A1A1A"
    C_LIVE    = "E8F5E9"   # green
    C_SCHED   = "FFF3E0"   # orange
    C_PLANNED = "F5F5F5"   # grey
    C_HIGH    = "FFF9C4"   # yellow accent for priority
    C_GROUP   = {
        "Sporting Group":     "E3F2FD",
        "Hound Group":        "FCE4EC",
        "Working Group":      "EDE7F6",
        "Terrier Group":      "E8F5E9",
        "Toy Group":          "FFF3E0",
        "Non-Sporting Group": "F3E5F5",
        "Herding Group":      "E0F7FA",
    }

    headers1 = ["#", "Breed Name", "AKC Group", "Size", "Priority",
                 "Status", "Main Guide", "Grooming Guide", "First-Year Costs", "Puppy Checklist"]
    widths1  = [4, 32, 22, 16, 10, 14, 52, 52, 52, 52]

    ws1.row_dimensions[1].height = 26
    for col, (h, w) in enumerate(zip(headers1, widths1), 1):
        hcell(ws1, 1, col, h)
        ws1.column_dimensions[get_column_letter(col)].width = w

    row = 2
    for i, (name, group, size, priority) in enumerate(BREEDS, 1):
        slug  = name_to_slug(name)
        if slug in existing_slugs:
            status = "Scheduled" if slug_status.get(slug) == "scheduled" else "Live"
        else:
            status = "Planned"

        s_bg = C_LIVE if status == "Live" else (C_SCHED if status == "Scheduled" else C_PLANNED)
        g_bg = C_GROUP.get(group, "FFFFFF")

        main_url     = f"{BLOG_BASE}/{slug}"
        groom_url    = f"{BLOG_BASE}/{slug}-grooming-guide"
        costs_url    = f"{BLOG_BASE}/{slug}-first-year-costs"
        checklist_url= f"{BLOG_BASE}/{slug}-puppy-checklist"

        ws1.row_dimensions[row].height = 18
        dcell(ws1, row, 1,  i,         bg=g_bg, align="center")
        dcell(ws1, row, 2,  name,      bg=g_bg, bold=(priority=="High"))
        dcell(ws1, row, 3,  group,     bg=g_bg)
        dcell(ws1, row, 4,  size,      bg=g_bg)
        dcell(ws1, row, 5,  priority,  bg=C_HIGH if priority=="High" else g_bg, align="center",
              bold=(priority=="High"), color=("B45309" if priority=="High" else "1A1A1A"))
        dcell(ws1, row, 6,  status,    bg=s_bg, align="center",
              bold=(status != "Planned"), color=("1B5E20" if status=="Live" else ("E65100" if status=="Scheduled" else "757575")))
        dcell(ws1, row, 7,  main_url,      bg=s_bg if status != "Planned" else "FFFFFF", color="0563C1" if status!="Planned" else "9E9E9E", hyperlink=main_url if status!="Planned" else None)
        dcell(ws1, row, 8,  groom_url,     bg="FFFFFF", color="0563C1" if status!="Planned" else "9E9E9E", hyperlink=groom_url if status!="Planned" else None)
        dcell(ws1, row, 9,  costs_url,     bg="FFFFFF", color="0563C1" if status!="Planned" else "9E9E9E", hyperlink=costs_url if status!="Planned" else None)
        dcell(ws1, row, 10, checklist_url, bg="FFFFFF", color="0563C1" if status!="Planned" else "9E9E9E", hyperlink=checklist_url if status!="Planned" else None)

        row += 1

    ws1.freeze_panes    = "A2"
    ws1.auto_filter.ref = f"A1:J{row-1}"

    # ── Sheet 2: Summary by Group ─────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.column_dimensions["A"].width = 24
    ws2.column_dimensions["B"].width = 10
    ws2.column_dimensions["C"].width = 10
    ws2.column_dimensions["D"].width = 10
    ws2.column_dimensions["E"].width = 12
    ws2.column_dimensions["F"].width = 12

    for col, h in enumerate(["AKC Group", "Total", "Live", "Scheduled", "Planned", "% Done"], 1):
        hcell(ws2, 1, col, h)
    ws2.row_dimensions[1].height = 26

    existing_slugs_set = existing_slugs
    slug_status_map    = slug_status

    group_names = list(dict.fromkeys(b[1] for b in BREEDS))
    srow = 2
    grand = {"total": 0, "live": 0, "sched": 0, "planned": 0}
    for g in group_names:
        breeds_in_group = [b for b in BREEDS if b[1] == g]
        total   = len(breeds_in_group)
        live    = sum(1 for b in breeds_in_group if name_to_slug(b[0]) in existing_slugs_set and slug_status_map.get(name_to_slug(b[0])) != "scheduled")
        sched   = sum(1 for b in breeds_in_group if name_to_slug(b[0]) in existing_slugs_set and slug_status_map.get(name_to_slug(b[0])) == "scheduled")
        planned = total - live - sched
        pct     = f"{round((live+sched)/total*100)}%" if total else "0%"
        g_bg    = C_GROUP.get(g, "FFFFFF").lstrip("#") if C_GROUP.get(g, "FFFFFF").startswith("#") else C_GROUP.get(g, "FFFFFF")
        for col, val in enumerate([g, total, live, sched, planned, pct], 1):
            dcell(ws2, srow, col, val, bg=g_bg, align="center" if col > 1 else "left")
        grand["total"]   += total
        grand["live"]    += live
        grand["sched"]   += sched
        grand["planned"] += planned
        srow += 1

    # totals row
    total_pct = f"{round((grand['live']+grand['sched'])/grand['total']*100)}%" if grand["total"] else "0%"
    for col, val in enumerate(["TOTAL", grand["total"], grand["live"], grand["sched"], grand["planned"], total_pct], 1):
        hcell(ws2, srow, col, val, bg="1A1A1A", fg="FFFFFF")

    wb.save(OUT_PATH)
    live_count  = sum(1 for b in BREEDS if name_to_slug(b[0]) in existing_slugs_set and slug_status_map.get(name_to_slug(b[0])) != "scheduled")
    sched_count = sum(1 for b in BREEDS if name_to_slug(b[0]) in existing_slugs_set and slug_status_map.get(name_to_slug(b[0])) == "scheduled")
    print(f"Saved {OUT_PATH}")
    print(f"  Total breeds: {len(BREEDS)}")
    print(f"  Live: {live_count}  |  Scheduled: {sched_count}  |  Planned: {len(BREEDS)-live_count-sched_count}")


if __name__ == "__main__":
    build_plan()
